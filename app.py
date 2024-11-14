import mysql.connector
import os
import re
import tempfile
from openpyxl import Workbook
from Crypto.Hash import MD4
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.urls import url_quote
from ldap3 import Server, Connection, ALL, NTLM, SUBTREE, MODIFY_REPLACE
from ldap3.core.exceptions import LDAPBindError, LDAPSocketOpenError, LDAPInvalidCredentialsResult
from datetime import datetime, timedelta
from mysql.connector import Error
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Configure MySQL connection using mysql-connector-python
db_config = {
    'host': os.getenv('DB_HOST'),
    'port': int(os.getenv('DB_PORT')),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'database': os.getenv('DB_NAME'),
    'auth_plugin': 'mysql_native_password',
}

def get_db_connection():
    """Establish a MySQL database connection."""
    return mysql.connector.connect(**db_config)


def authenticate_ad(username, password, domain, ad_server):
    try:
        server = Server(ad_server, get_info=ALL)
        user_dn = f"{domain}\\{username}"
        conn = Connection(server, user=user_dn, password=password, authentication=NTLM, auto_bind=True)
        if conn.bound:
            return conn
        return None

    except (LDAPInvalidCredentialsResult, LDAPSocketOpenError, LDAPBindError):
        return None

def convert_windows_timestamp(timestamp):
    if not timestamp:
        return None
    if isinstance(timestamp, datetime):
        return timestamp
    windows_epoch = datetime(1601, 1, 1)
    return windows_epoch + timedelta(seconds=int(timestamp) / 10**7)

def get_user_status(user_account_control):
    return 'Disabled' if int(user_account_control) & 2 else 'Enabled'

def get_all_users(conn, base_dn):
    conn.search(
        search_base=base_dn,
        search_filter='(&(objectClass=user)(!(sAMAccountName=*\\$)))',
        search_scope=SUBTREE,
        attributes=['sAMAccountName', 'displayName', 'memberOf', 'lastLogonTimestamp', 'pwdLastSet', 'userAccountControl']
    )

    users = []
    for entry in conn.entries:
        users.append({
            "username": entry.sAMAccountName.value,
            "display_name": entry.displayName.value,
            "member_of": ";".join(entry.memberOf.values) if entry.memberOf else "",
            "last_login": convert_windows_timestamp(entry.lastLogonTimestamp.value),
            "last_password_change": convert_windows_timestamp(entry.pwdLastSet.value),
            "status": get_user_status(entry.userAccountControl.value),
        })
    return users

# Password strength check
def is_password_strong(password):
    if (len(password) >= 8 and 
        re.search(r"[A-Za-z]", password) and
        re.search(r"\d", password) and
        re.search(r"[!@#$%^&*()_+]", password)):
        return True
    return False

# Summary Domain Retrieval
def summary():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT * FROM ad_summary')
        if cursor.fetchone()[0]  == 0:
            return
        
        summaries = [
                    ("sum_admin", "SELECT COUNT(id) FROM ad_users WHERE member_of LIKE %s", ('%Domain Admins%',)),
                    ("sum_disabled", "SELECT COUNT(id) FROM ad_users WHERE status = %s AND tipe='user'", ('Disabled',)),
                    ("sum_notused", "SELECT COUNT(id) FROM ad_users WHERE last_login <= CURDATE() - INTERVAL 3 MONTH AND tipe='user'", ()),
                    ("sum_outoffdate_pass", "SELECT COUNT(id) FROM ad_users WHERE tipe='user' AND status='Enabled' AND last_login >= CURDATE() - INTERVAL 1 MONTH AND last_password_change <= CURDATE() - INTERVAL 2 MONTH", ()),
                    ("sum_computer", "SELECT COUNT(id) FROM ad_users WHERE tipe = %s", ('komputer',)),
                    ("sum_user", "SELECT COUNT(id) FROM ad_users WHERE tipe = %s", ('user',)),
                    ("sum_disabled_computer","SELECT COUNT(id) FROM ad_users WHERE status = %s AND tipe='komputer'",('Disabled',))
                ]
    
        # Update each summary value in the ad_summary table
        for column, query, params in summaries:
                    cursor.execute(query, params)
                    result = cursor.fetchone()[0]
                    cursor.execute(f"UPDATE ad_summary SET {column}=%s", (result,))
                
        conn.commit()  # Commit all updates at once after calculations
            
    except Exception as e:
        # Rollback and log the error without exposing sensitive details
        conn.rollback()
        flash("An error occurred while updating the summary.", 'error')
        print(f"Error details: {e}")  # Log the actual error for debugging
        
# Login Route
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if user and check_password_hash(user[2], password):
            session['user_id'] = user[0]
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password!','danger')
    return render_template('login.html')

# Dashboard Route
@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()

    if user:
        username = user[0]
    else:
        flash('User not found.', 'error')
        return redirect(url_for('login'))
    
    return render_template('dashboard.html', username=username)

# User Manager Route
@app.route('/user_manager', methods=['GET', 'POST'])
def user_manager():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        new_username = request.form['username']
        new_password = request.form['password']
        
        cursor.execute("SELECT * FROM users WHERE username = %s", (new_username,))
        existing_user = cursor.fetchone()
        
        if existing_user:
            flash('Username already exists, please choose another.', 'error')
        elif not is_password_strong(new_password):
            flash('Password must be at least 8 characters, with letters, numbers, and special characters.', 'error')
        else:
            hashed_password = generate_password_hash(new_password)
            cursor.execute("INSERT INTO users (username, password) VALUES (%s, %s)", (new_username, hashed_password))
            conn.commit()
            flash('User successfully registered!', 'success')
    
    cursor.execute("SELECT username FROM users")
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('user_manager.html', users=users)

@app.route('/delete_user/<username>', methods=['POST'])
def delete_user(username):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM users WHERE username = %s", (username,))
    conn.commit()
    cursor.close()
    conn.close()

    flash(f'User {username} has been deleted.', 'success')
    
    return redirect(url_for('user_manager'))

# AD Auditing Route
@app.route('/ad_auditing', methods=['GET', 'POST'])
def ad_auditing():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get the search term from the query parameters
    search_term = request.args.get('search', '').strip()
    stored_users = []  # Initialize stored_users to ensure it's always defined
    users = []
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        domain = request.form['domain']
        ad_server = request.form['ad_server']
        base_dn = request.form['base_dn']

        conn = authenticate_ad(username, password, domain, ad_server)
        if conn:
            users = get_all_users(conn, base_dn)
            conn.unbind()

            dbconn = get_db_connection()
            cursor = dbconn.cursor()
            # Simpan pengguna ke tabel MySQL

            # Kosongkan tabel ad_users sebelum menyimpan data baru
            cursor.execute("DELETE FROM ad_users")
            
            # Simpan data ke database
            for user in users:
                user_type = 'komputer' if '$' in user["username"] else 'user'
                cursor.execute(
                    "INSERT INTO ad_users (username, display_name, last_login, status, domainname, tipe, member_of, last_password_change) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                    (user["username"], user["display_name"], user["last_login"], user["status"], domain, user_type, user["member_of"], user["last_password_change"])
                )
            
            dbconn.commit()
            
            # Kosongkan table ad_summary
            cursor.execute("DELETE FROM ad_summary")
            
            # Simpan intial data ke tabel ad_summary
            cursor.execute(
                "INSERT INTO ad_summary (ad_name, sum_admin, sum_disabled, sum_notused, sum_outoffdate_pass, sum_computer, sum_user, sum_disabled_computer) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                (domain, 0, 0, 0, 0, 0, 0, 0)
            )
            
            dbconn.commit()
            cursor.close()
            
            summary()
            print(f'Successfully imported {len(users)} record from Active Directory!')
        else:
            print('Failed retrieve users from Active Directory!')

    # Ambil data dari tabel ad_users setelah penyimpanan
    try:
        dbconn = get_db_connection()
        cursor = dbconn.cursor(dictionary=True)
        if search_term:
            query = """
                SELECT * FROM ad_users 
                WHERE username LIKE %s OR display_name LIKE %s OR member_of LIKE %s 
                ORDER BY username ASC
            """    
            cursor.execute(query, ('%' + search_term + '%', '%' + search_term + '%', '%' + search_term + '%'))
        else:   
            query = "SELECT * FROM ad_users ORDER BY username ASC"         
            cursor.execute(query)
            
        # Fetch results after the query
        stored_users = cursor.fetchall()
        
    except Error as e:
            flash(f"Error: {e}", 'error')
            stored_users = []  # Empty list if there's an error
    finally:
        # Ensure the cursor and database connection are closed properly
        if cursor:
            cursor.close()
        if dbconn:
            dbconn.close()

    return render_template('ad_auditing.html', users=stored_users)

@app.route('/export_summary', methods=['GET'])
def export_summary():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Query 1: General summary data
        cursor.execute("SELECT ad_name as domain, sum_admin as domain_admin, sum_disabled as disabled_user, sum_disabled_computer as disabled_computer, sum_notused as user_lama_tidak_login, sum_outoffdate_pass as user_aktif_pass_tdk_ganti, sum_computer as jmlh_komputer, sum_user as jmlh_user FROM ad_summary")
        summary_data = cursor.fetchall()
        summary_columns = [desc[0] for desc in cursor.description]

        # Query 2: All domain admins
        cursor.execute("SELECT domainname, username, display_name, member_of, last_login, last_password_change FROM ad_users WHERE member_of LIKE '%Domain Admins%'")
        admins_data = cursor.fetchall()
        admins_columns = [desc[0] for desc in cursor.description]

        # Query 3: All disabled users
        cursor.execute("SELECT domainname, username, display_name, member_of, last_login, last_password_change FROM ad_users WHERE status LIKE 'Disabled' AND tipe='user'")
        disabled_data = cursor.fetchall()
        disabled_columns = [desc[0] for desc in cursor.description]

        # Query 4: Inactive users (last login > 3 months)
        cursor.execute("SELECT domainname, username, display_name, member_of, last_login, last_password_change FROM ad_users WHERE tipe='user' AND last_login <= CURDATE() - INTERVAL 3 MONTH")
        inactive_data = cursor.fetchall()
        inactive_columns = [desc[0] for desc in cursor.description]
         
        # Query 5: Active users with outdated passwords
        cursor.execute("SELECT domainname, username, display_name, member_of, last_login, last_password_change FROM ad_users WHERE tipe='user' AND status='Enabled' AND last_login >= ( SELECT CURDATE() - INTERVAL 1 MONTH ) AND last_password_change <= ( CURDATE() - INTERVAL 2 MONTH )")
        user_expired_data = cursor.fetchall()
        user_expired_columns = [desc[0] for desc in cursor.description]
        
        # Query 6: Disabled Computer
        cursor.execute("SELECT domainname, username, last_login FROM ad_users WHERE tipe='komputer' AND status='Disabled'")
        disabled_computer_data = cursor.fetchall()
        disabled_computer_columns = [desc[0] for desc in cursor.description]
    
    finally:
        cursor.close()
        conn.close()
        
    # Create Excel workbook and add data to multiple sheets
    workbook = Workbook()

    # Function to add data to a worksheet
    def add_data_to_sheet(sheet, columns, data):
        sheet.append(columns)  # Add column headers
        for row in data:
            sheet.append(row)  # Add data rows

    # Sheet 1: Summary
    sheet_summary = workbook.active
    sheet_summary.title = "Summary"
    add_data_to_sheet(sheet_summary, summary_columns, summary_data)

    # Sheet 2: Domain Admins
    sheet_admins = workbook.create_sheet(title="Domain Admins")
    add_data_to_sheet(sheet_admins, admins_columns, admins_data)

    # Sheet 3: Disabled Users
    sheet_disabled = workbook.create_sheet(title="Disabled Users")
    add_data_to_sheet(sheet_disabled, disabled_columns, disabled_data)

    # Sheet 4: Inactive Users
    sheet_inactive = workbook.create_sheet(title="User Lama Tidak Login")
    add_data_to_sheet(sheet_inactive, inactive_columns, inactive_data)

    # Sheet 5: Active Users with Outdated Passwords
    sheet_user_expired = workbook.create_sheet(title="User Aktif Pswd Tdk Ganti")
    add_data_to_sheet(sheet_user_expired, user_expired_columns, user_expired_data)
    
    # Sheet 6: Disabled Computer
    sheet_computer_disabled = workbook.create_sheet(title="Disabled Computers")
    add_data_to_sheet(sheet_computer_disabled, disabled_computer_columns, disabled_computer_data)

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_path = tmp.name
        temp_file.close()  # Close the file so `send_file` can access it
        workbook.save(excel_path)
        
    return send_file(
        excel_path,
        as_attachment=True,
        download_name="ad_summary_multisheet_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Logout Route
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
